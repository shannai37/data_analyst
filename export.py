"""
数据分析师插件 - 数据导出模块

提供多种格式的数据导出功能
"""

import json
import time
import os
import aiosqlite
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any

# Excel处理
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, LineChart, Reference

# PDF生成
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from astrbot.api import logger
from .models import ExportConstants as EC, PluginConfig
from .database import DatabaseManager


class ExportManager:
    """
    /// 数据导出管理器
    /// 支持Excel、PDF、CSV、JSON等多种格式的数据导出
    /// 提供专业的报告生成和数据可视化
    """
    
    def __init__(self, exports_dir: Path, db_manager: DatabaseManager, config: PluginConfig):
        """
        /// 初始化导出管理器
        /// @param exports_dir: 导出文件目录
        /// @param db_manager: 数据库管理器
        /// @param config: 插件配置
        """
        self.exports_dir = exports_dir
        self.db_manager = db_manager
        self.config = config
        self.exports_dir.mkdir(exist_ok=True)
        
        # 初始化PDF字体
        self._setup_pdf_fonts()
        
        logger.info(f"数据导出管理器已初始化: {exports_dir}")
    
    def _setup_pdf_fonts(self):
        """设置PDF中文字体支持"""
        try:
            # 注册中文字体（如果可用）
            # 这里使用系统默认字体，实际部署时可能需要包含字体文件
            pass
        except Exception as e:
            logger.warning(f"PDF字体设置失败: {e}")
    
    async def export_to_excel(self, group_id: str, period: str) -> Optional[str]:
        """
        /// 导出Excel格式报告
        /// @param group_id: 群组ID
        /// @param period: 时间周期
        /// @return: Excel文件路径
        """
        try:
            # 获取数据
            activity_data = await self.db_manager.get_activity_analysis(group_id, period)
            topics_data = await self.db_manager.get_topics_analysis(group_id, period)
            group_stats = await self.db_manager.get_group_quick_stats(group_id)
            
            # 生成文件名
            timestamp = int(time.time())
            filename = EC.EXCEL_TEMPLATE.format(
                group_id=group_id, period=period, timestamp=timestamp
            )
            filepath = self.exports_dir / filename
            
            # 创建Excel工作簿
            with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
                # 分析摘要工作表
                self._create_summary_sheet(writer, group_stats, activity_data, topics_data, period)
                
                # 活跃度统计工作表
                if activity_data and activity_data.daily_data:
                    self._create_activity_sheet(writer, activity_data)
                
                # 热门话题工作表
                if topics_data and topics_data.top_topics:
                    self._create_topics_sheet(writer, topics_data)
                
                # 用户排行榜工作表
                await self._create_user_ranking_sheet(writer, group_id, period)
            
            # 美化Excel格式
            await self._format_excel_workbook(filepath)
            
            logger.info(f"Excel报告导出成功: {filename}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"Excel导出失败: {e}")
            return None
    
    def _create_summary_sheet(self, writer, group_stats: Dict, activity_data, topics_data, period: str):
        """创建分析摘要工作表"""
        summary_data = []
        
        # 基础信息
        summary_data.append(['报告生成时间', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        summary_data.append(['分析周期', period])
        summary_data.append(['', ''])  # 空行
        
        # 群组基础统计
        if group_stats:
            summary_data.append(['=== 群组基础统计 ===', ''])
            summary_data.append(['总消息数', group_stats.get('total_messages', 0)])
            summary_data.append(['活跃用户数', group_stats.get('active_users', 0)])
            summary_data.append(['平均消息长度', f"{group_stats.get('avg_message_length', 0):.1f}字"])
            summary_data.append(['数据收集天数', group_stats.get('data_days', 0)])
            summary_data.append(['最活跃时段', f"{group_stats.get('peak_hour', 'N/A')}时"])
        
        summary_data.append(['', ''])  # 空行
        
        # 活跃度分析
        if activity_data:
            summary_data.append(['=== 活跃度分析 ===', ''])
            summary_data.append(['周期内总消息', activity_data.total_messages])
            summary_data.append(['周期内活跃用户', activity_data.active_users])
            summary_data.append(['日均消息数', f"{activity_data.avg_daily_messages:.1f}"])
            summary_data.append(['增长率', f"{activity_data.growth_rate:+.1f}%"])
            summary_data.append(['趋势描述', activity_data.trend_description])
        
        summary_data.append(['', ''])  # 空行
        
        # 话题分析
        if topics_data:
            summary_data.append(['=== 话题分析 ===', ''])
            summary_data.append(['热门话题数量', len(topics_data.top_topics)])
            summary_data.append(['新话题数量', topics_data.new_topics_count])
            summary_data.append(['话题活跃度', f"{topics_data.topic_activity:.1f}%"])
            summary_data.append(['讨论深度', f"{topics_data.discussion_depth:.1f}次/话题"])
        
        # 创建DataFrame并写入
        df = pd.DataFrame(summary_data, columns=['指标', '数值'])
        df.to_excel(writer, sheet_name=EC.SHEET_SUMMARY, index=False)
    
    def _create_activity_sheet(self, writer, activity_data):
        """创建活跃度统计工作表"""
        # 每日活跃度数据
        daily_df = pd.DataFrame(activity_data.daily_data, columns=['日期', '消息数'])
        daily_df['日期'] = pd.to_datetime(daily_df['日期'])
        daily_df['星期'] = daily_df['日期'].dt.day_name()
        daily_df['累计消息数'] = daily_df['消息数'].cumsum()
        
        daily_df.to_excel(writer, sheet_name=EC.SHEET_ACTIVITY, index=False)
    
    def _create_topics_sheet(self, writer, topics_data):
        """创建热门话题工作表"""
        topics_df = pd.DataFrame(topics_data.top_topics)
        topics_df['排名'] = range(1, len(topics_df) + 1)
        topics_df['最后提及时间'] = pd.to_datetime(topics_df['last_mentioned'])
        
        # 重新排列列顺序
        topics_df = topics_df[['排名', 'keyword', 'frequency', '最后提及时间']]
        topics_df.columns = ['排名', '关键词', '频次', '最后提及时间']
        
        topics_df.to_excel(writer, sheet_name=EC.SHEET_TOPICS, index=False)
    
    async def _create_user_ranking_sheet(self, writer, group_id: str, period: str):
        """创建用户排行榜工作表"""
        try:
            # 获取用户排行数据
            async with aiosqlite.connect(self.db_manager.db_path) as db:
                cursor = await db.execute('''
                    SELECT user_id, COUNT(*) as message_count, 
                           SUM(word_count) as total_words,
                           AVG(word_count) as avg_words
                    FROM messages 
                    WHERE group_id = ? AND timestamp >= ?
                    GROUP BY user_id
                    ORDER BY message_count DESC
                    LIMIT 50
                ''', (group_id, self.db_manager._calculate_start_date(period)))
                
                user_data = await cursor.fetchall()
                
                if user_data:
                    df = pd.DataFrame(user_data, columns=['用户ID', '消息数', '总字数', '平均字数'])
                    df['排名'] = range(1, len(df) + 1)
                    df = df[['排名', '用户ID', '消息数', '总字数', '平均字数']]
                    df['平均字数'] = df['平均字数'].round(1)
                    
                    df.to_excel(writer, sheet_name=EC.SHEET_USER_RANKING, index=False)
                    
        except Exception as e:
            logger.error(f"用户排行榜数据获取失败: {e}")
    
    async def _format_excel_workbook(self, filepath: str):
        """美化Excel工作簿格式"""
        try:
            wb = openpyxl.load_workbook(filepath)
            
            # 定义样式
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # 格式化每个工作表
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # 设置表头样式
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = Alignment(horizontal='center')
                        cell.border = border
                
                # 设置数据行样式
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    for cell in row:
                        cell.border = border
                        cell.alignment = Alignment(horizontal='left')
                
                # 自动调整列宽
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(filepath)
            
        except Exception as e:
            logger.error(f"Excel格式化失败: {e}")
    
    async def export_to_pdf(self, group_id: str, period: str) -> Optional[str]:
        """
        /// 导出PDF格式报告
        /// @param group_id: 群组ID
        /// @param period: 时间周期
        /// @return: PDF文件路径
        """
        try:
            # 获取数据
            activity_data = await self.db_manager.get_activity_analysis(group_id, period)
            topics_data = await self.db_manager.get_topics_analysis(group_id, period)
            group_stats = await self.db_manager.get_group_quick_stats(group_id)
            
            # 生成文件名
            timestamp = int(time.time())
            filename = EC.PDF_TEMPLATE.format(
                group_id=group_id, period=period, timestamp=timestamp
            )
            filepath = self.exports_dir / filename
            
            # 创建PDF文档
            doc = SimpleDocTemplate(str(filepath), pagesize=A4)
            styles = getSampleStyleSheet()
            story = []
            
            # 添加标题
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=18,
                spaceAfter=30,
                alignment=1  # 居中
            )
            story.append(Paragraph(EC.REPORT_TITLE, title_style))
            story.append(Spacer(1, 12))
            
            # 添加基础信息
            story.append(Paragraph(f"<b>报告生成时间:</b> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal']))
            story.append(Paragraph(f"<b>分析周期:</b> {period}", styles['Normal']))
            story.append(Paragraph(f"<b>群组ID:</b> {group_id}", styles['Normal']))
            story.append(Spacer(1, 20))
            
            # 添加摘要信息
            if group_stats:
                story.append(Paragraph("<b>群组统计摘要</b>", styles['Heading2']))
                summary_data = [
                    ['指标', '数值'],
                    ['总消息数', str(group_stats.get('total_messages', 0))],
                    ['活跃用户数', str(group_stats.get('active_users', 0))],
                    ['平均消息长度', f"{group_stats.get('avg_message_length', 0):.1f}字"],
                    ['数据收集天数', str(group_stats.get('data_days', 0))],
                    ['最活跃时段', f"{group_stats.get('peak_hour', 'N/A')}时"]
                ]
                
                summary_table = Table(summary_data)
                summary_table.setStyle(self._get_table_style())
                story.append(summary_table)
                story.append(Spacer(1, 20))
            
            # 添加活跃度分析
            if activity_data:
                story.append(Paragraph("<b>活跃度分析</b>", styles['Heading2']))
                activity_content = f"""
                <b>周期内总消息:</b> {activity_data.total_messages}<br/>
                <b>活跃用户数:</b> {activity_data.active_users}<br/>
                <b>日均消息数:</b> {activity_data.avg_daily_messages:.1f}<br/>
                <b>增长率:</b> {activity_data.growth_rate:+.1f}%<br/>
                <b>趋势描述:</b> {activity_data.trend_description}
                """
                story.append(Paragraph(activity_content, styles['Normal']))
                story.append(Spacer(1, 20))
            
            # 添加话题分析
            if topics_data and topics_data.top_topics:
                story.append(Paragraph("<b>热门话题分析</b>", styles['Heading2']))
                
                # 创建话题表格
                topics_table_data = [['排名', '关键词', '频次']]
                for i, topic in enumerate(topics_data.top_topics[:10], 1):
                    topics_table_data.append([
                        str(i),
                        topic['keyword'],
                        str(topic['frequency'])
                    ])
                
                topics_table = Table(topics_table_data)
                topics_table.setStyle(self._get_table_style())
                story.append(topics_table)
                story.append(Spacer(1, 20))
            
            # 添加页脚信息
            story.append(Spacer(1, 30))
            story.append(Paragraph("--- 报告结束 ---", styles['Normal']))
            story.append(Paragraph(f"由 AstrBot 数据分析师插件生成", styles['Normal']))
            
            # 生成PDF
            doc.build(story)
            
            logger.info(f"PDF报告导出成功: {filename}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"PDF导出失败: {e}")
            return None
    
    def _get_table_style(self):
        """获取表格样式"""
        return TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
    
    async def export_to_csv(self, group_id: str, period: str) -> Optional[str]:
        """
        /// 导出CSV格式数据
        /// @param group_id: 群组ID
        /// @param period: 时间周期
        /// @return: CSV文件路径
        """
        try:
            # 获取原始数据
            start_date = self.db_manager._calculate_start_date(period)
            
            async with aiosqlite.connect(self.db_manager.db_path) as db:
                cursor = await db.execute('''
                    SELECT message_id, user_id, group_id, platform, message_type,
                           timestamp, word_count, created_at
                    FROM messages 
                    WHERE group_id = ? AND timestamp >= ?
                    ORDER BY timestamp
                    LIMIT ?
                ''', (group_id, start_date, EC.MAX_EXPORT_ROWS))
                
                data = await cursor.fetchall()
                
                if not data:
                    return None
                
                # 创建DataFrame
                df = pd.DataFrame(data, columns=[
                    '消息ID', '用户ID', '群组ID', '平台', '消息类型',
                    '时间戳', '字数', '创建时间'
                ])
                
                # 生成文件名
                timestamp = int(time.time())
                filename = EC.CSV_TEMPLATE.format(
                    group_id=group_id, period=period, timestamp=timestamp
                )
                filepath = self.exports_dir / filename
                
                # 导出CSV
                df.to_csv(filepath, index=False, encoding='utf-8-sig')
                
                logger.info(f"CSV数据导出成功: {filename}, 记录数: {len(data)}")
                return str(filepath)
                
        except Exception as e:
            logger.error(f"CSV导出失败: {e}")
            return None
    
    async def export_to_json(self, group_id: str, period: str) -> Optional[str]:
        """
        /// 导出JSON格式数据
        /// @param group_id: 群组ID
        /// @param period: 时间周期
        /// @return: JSON文件路径
        """
        try:
            # 获取综合分析数据
            activity_data = await self.db_manager.get_activity_analysis(group_id, period)
            topics_data = await self.db_manager.get_topics_analysis(group_id, period)
            group_stats = await self.db_manager.get_group_quick_stats(group_id)
            
            # 构建JSON数据结构
            export_data = {
                'export_info': {
                    'group_id': group_id,
                    'period': period,
                    'export_time': datetime.now().isoformat(),
                    'export_version': '1.0'
                },
                'group_stats': group_stats or {},
                'activity_analysis': activity_data.to_dict() if activity_data else {},
                'topics_analysis': topics_data.to_dict() if topics_data else {}
            }
            
            # 生成文件名
            timestamp = int(time.time())
            filename = EC.JSON_TEMPLATE.format(
                group_id=group_id, period=period, timestamp=timestamp
            )
            filepath = self.exports_dir / filename
            
            # 导出JSON
            with open(filepath, 'w', encoding='utf-8') as f:
                json.dump(export_data, f, ensure_ascii=False, indent=2)
            
            logger.info(f"JSON数据导出成功: {filename}")
            return str(filepath)
            
        except Exception as e:
            logger.error(f"JSON导出失败: {e}")
            return None
    
    async def create_comprehensive_report(self, group_id: str, period: str, 
                                        include_charts: bool = True) -> Optional[str]:
        """
        /// 创建综合分析报告
        /// @param group_id: 群组ID
        /// @param period: 时间周期
        /// @param include_charts: 是否包含图表
        /// @return: 报告文件路径
        """
        try:
            # 创建包含所有格式的综合报告目录
            timestamp = int(time.time())
            report_dir = self.exports_dir / f"comprehensive_report_{group_id}_{timestamp}"
            report_dir.mkdir(exist_ok=True)
            
            # 导出各种格式
            results = {}
            
            # Excel报告
            excel_path = await self.export_to_excel(group_id, period)
            if excel_path:
                new_excel_path = report_dir / "analysis_report.xlsx"
                os.rename(excel_path, new_excel_path)
                results['excel'] = str(new_excel_path)
            
            # PDF报告
            pdf_path = await self.export_to_pdf(group_id, period)
            if pdf_path:
                new_pdf_path = report_dir / "analysis_report.pdf"
                os.rename(pdf_path, new_pdf_path)
                results['pdf'] = str(new_pdf_path)
            
            # CSV数据
            csv_path = await self.export_to_csv(group_id, period)
            if csv_path:
                new_csv_path = report_dir / "raw_data.csv"
                os.rename(csv_path, new_csv_path)
                results['csv'] = str(new_csv_path)
            
            # JSON数据
            json_path = await self.export_to_json(group_id, period)
            if json_path:
                new_json_path = report_dir / "analysis_data.json"
                os.rename(json_path, new_json_path)
                results['json'] = str(new_json_path)
            
            # 创建报告说明文件
            readme_content = f"""# 群组数据分析综合报告

## 报告信息
- 群组ID: {group_id}
- 分析周期: {period}
- 生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}
- 报告版本: 1.0

## 文件说明
- analysis_report.xlsx: Excel格式的详细分析报告
- analysis_report.pdf: PDF格式的可视化报告
- raw_data.csv: 原始消息数据（CSV格式）
- analysis_data.json: 分析结果数据（JSON格式）

## 使用说明
1. Excel文件包含多个工作表，分别展示不同维度的分析结果
2. PDF文件提供了图文并茂的分析报告，适合演示和分享
3. CSV文件包含原始数据，可用于进一步的自定义分析
4. JSON文件包含结构化的分析结果，便于程序化处理

## 技术支持
如有问题请联系管理员或查看插件文档。
"""
            
            readme_path = report_dir / "README.md"
            with open(readme_path, 'w', encoding='utf-8') as f:
                f.write(readme_content)
            
            logger.info(f"综合报告创建完成: {report_dir}")
            return str(report_dir)
            
        except Exception as e:
            logger.error(f"综合报告创建失败: {e}")
            return None
    
    async def cleanup_old_exports(self, max_age_days: int = 7):
        """
        /// 清理过期的导出文件
        /// @param max_age_days: 文件最大保留天数
        """
        try:
            current_time = time.time()
            max_age_seconds = max_age_days * 24 * 3600
            
            deleted_count = 0
            for export_file in self.exports_dir.rglob("*"):
                if export_file.is_file():
                    file_age = current_time - export_file.stat().st_mtime
                    if file_age > max_age_seconds:
                        export_file.unlink()
                        deleted_count += 1
            
            # 清理空目录
            for export_dir in self.exports_dir.iterdir():
                if export_dir.is_dir() and not any(export_dir.iterdir()):
                    export_dir.rmdir()
            
            if deleted_count > 0:
                logger.info(f"已清理 {deleted_count} 个过期导出文件")
                
        except Exception as e:
            logger.error(f"导出文件清理失败: {e}")
    
    def get_export_stats(self) -> Dict[str, Any]:
        """
        /// 获取导出统计信息
        /// @return: 统计数据
        """
        try:
            export_files = list(self.exports_dir.rglob("*"))
            file_count_by_type = {}
            total_size = 0
            
            for file_path in export_files:
                if file_path.is_file():
                    suffix = file_path.suffix.lower()
                    file_count_by_type[suffix] = file_count_by_type.get(suffix, 0) + 1
                    total_size += file_path.stat().st_size
            
            return {
                'total_files': len([f for f in export_files if f.is_file()]),
                'total_size_mb': total_size / (1024 * 1024),
                'file_count_by_type': file_count_by_type,
                'exports_dir': str(self.exports_dir)
            }
            
        except Exception as e:
            logger.error(f"获取导出统计失败: {e}")
            return {}
